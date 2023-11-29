import os
from datetime import datetime

# Imports from PyTorch.
from torch import nn, Tensor, device, no_grad, manual_seed, save
from torch import max as torch_max
from torch.utils.data import DataLoader
import torch.nn.functional as F
import openpyxl
from torchvision import datasets, transforms
import torchvision.transforms as transforms
import torchvision.datasets as datasets
from torch.utils.data import DataLoader
import torch.utils.checkpoint as checkpoint
import torch
import pandas as pd
from aihwkit.optim import AnalogSGD
import random
import numpy as np
# class Args:
#     model = "resnet18"
#     dataset = "cifar100"
#     epoch = 100
#     loss_coefficient = 0.3
#     feature_loss_coefficient = 0.03
#     dataset_path = "data"
#     autoaugment = False
#     temperature = 3.0
#     batchsize = 128
#     init_lr = 0.1
#     N_CLASSES = 100
#     autoaugment = False
# args = Args()
# print(args)
# CUDA가 사용 가능한 경우 USE_CUDA를 1로 설정
USE_CUDA = 0
if torch.cuda.is_available():
    USE_CUDA = 1

# DEVICE 정의
DEVICE = device("cuda" if USE_CUDA else "cpu")
def CrossEntropy(outputs, targets,args):
    log_softmax_outputs = F.log_softmax(outputs/args.temperature, dim=1)
    softmax_targets = F.softmax(targets/args.temperature, dim=1)
    return -(log_softmax_outputs * softmax_targets).sum(dim=1).mean()

import pandas as pd
from openpyxl import load_workbook

# 결과를 저장할 디렉토리 생성
results_dir = "results"
if not os.path.exists(results_dir):
    os.makedirs(results_dir)

filepath = f"results/results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None, **to_excel_kwargs):
    # Excel 파일에 데이터프레임 추가
    if not os.path.isfile(filename):
        df.to_excel(filename, sheet_name, index=False, **to_excel_kwargs)
    else:
        book = load_workbook(filename)
        writer = pd.ExcelWriter(filename, engine='openpyxl')
        writer.sheets = {ws.title: ws for ws in book.worksheets}  # 기존 시트 가져오기
        
        if startrow is None:
            startrow = book[sheet_name].max_row
        
        df.to_excel(writer, sheet_name, index=False, header=False, startrow=startrow)
        writer.save()

def set_seed(seed_value=42):
    """모든 난수 생성기에 대해 시드를 설정하여 코드가 재현 가능하도록 합니다."""
    random.seed(seed_value)  # Python의 내장 random 모듈에 대한 시드 설정
    np.random.seed(seed_value)  # Numpy 모듈에 대한 시드 설정
    torch.manual_seed(seed_value)  # CPU를 위한 시드 설정
    torch.cuda.manual_seed(seed_value)  # 모든 GPU를 위한 시드 설정
    torch.cuda.manual_seed_all(seed_value)  # 멀티-GPU를 위한 시드 설정 (있는 경우)
    torch.backends.cudnn.deterministic = True  # CUDA를 위한 결정론적 모드 설정
    torch.backends.cudnn.benchmark = False  # 내부적으로 사용되는 cudnn auto-tuner를 비활성화

def worker_init_fn(worker_id):
    """워커 초기화 함수로, DataLoader의 각 워커에 대한 초기 시드를 설정합니다."""
    random.seed(42 + worker_id)
    np.random.seed(42 + worker_id)

def load_images(args):
    """Load images for training and testing."""
    # Mean and std for normalization
    mean = torch.Tensor([0.4914, 0.4822, 0.4465])
    std = torch.Tensor([0.2023, 0.1994, 0.2010])

    # 데이터 변형 시드 설정
    torch.manual_seed(42)

    # Set up training transform
    if args.autoaugment:
        transform_train = transforms.Compose([
            transforms.RandomCrop(32, padding=4, fill=128),
            transforms.RandomHorizontalFlip(),
            CIFAR10Policy(),
            transforms.ToTensor(),
            Cutout(n_holes=1, length=16),
            transforms.Normalize(mean, std)
        ])
    else:
        transform_train = transforms.Compose([
            transforms.RandomCrop(32, padding=4, fill=128),
            transforms.RandomHorizontalFlip(),
            transforms.ToTensor(),
            transforms.Normalize(mean, std)
        ])

    # Set up testing transform
    transform_test = transforms.Compose([
        transforms.ToTensor(),
        transforms.Normalize(mean, std)
    ])

    # Load the dataset
    if args.dataset == "cifar100":
        trainset = datasets.CIFAR100(
            root=args.dataset_path,
            train=True,
            download=True,
            transform=transform_train
        )
        testset = datasets.CIFAR100(
            root=args.dataset_path,
            train=False,
            download=True,
            transform=transform_test
        )
    elif args.dataset == "cifar10":
        trainset = datasets.CIFAR10(
            root=args.dataset_path,
            train=True,
            download=True,
            transform=transform_train
        )
        testset = datasets.CIFAR10(
            root=args.dataset_path,
            train=False,
            download=True,
            transform=transform_test
        )

    # DataLoader 생성 시 worker_init_fn을 추가하여 각 워커의 시드를 고정합니다.
    trainloader = DataLoader(trainset, batch_size=args.batchsize, shuffle=True, num_workers=4, worker_init_fn=worker_init_fn)
    testloader = DataLoader(testset, batch_size=args.batchsize, shuffle=False, num_workers=4, worker_init_fn=worker_init_fn)

    return trainloader, testloader

def conv3x3(in_planes, out_planes, stride=1, groups=1, dilation=1):
    """3x3 convolution with padding"""
    return nn.Conv2d(in_planes, out_planes, kernel_size=3, stride=stride,
                     padding=dilation, groups=groups, bias=False, dilation=dilation)


def conv1x1(in_planes, out_planes, stride=1):
    """1x1 convolution"""
    return nn.Conv2d(in_planes, out_planes, kernel_size=1, stride=stride, bias=False)



def ScalaNet(channel_in, channel_out, size):
    return nn.Sequential(
        nn.Conv2d(channel_in, 128, kernel_size=1, stride=1),
        nn.BatchNorm2d(128),
        nn.ReLU(),
        nn.Conv2d(128, 128, kernel_size=size, stride=size),
        nn.BatchNorm2d(128),
        nn.ReLU(),
        nn.Conv2d(128, channel_out, kernel_size=1, stride=1),
        nn.BatchNorm2d(channel_out),
        nn.ReLU(),
        nn.AvgPool2d(4, 4)
        )

class SepConv(nn.Module):

        def __init__(self, channel_in, channel_out, kernel_size=3, stride=2, padding=1, affine=True):
            super(SepConv, self).__init__()
            self.op = nn.Sequential(
                nn.Conv2d(channel_in, channel_in, kernel_size=kernel_size, stride=stride, padding=padding, groups=channel_in, bias=False),
                nn.Conv2d(channel_in, channel_in, kernel_size=1, padding=0, bias=False),
                nn.BatchNorm2d(channel_in, affine=affine),
                nn.ReLU(inplace=False),
                nn.Conv2d(channel_in, channel_in, kernel_size=kernel_size, stride=1, padding=padding, groups=channel_in, bias=False),
                nn.Conv2d(channel_in, channel_out, kernel_size=1, padding=0, bias=False),
                nn.BatchNorm2d(channel_out, affine=affine),
                nn.ReLU(inplace=False),
            )

        def forward(self, x):
            return self.op(x)

def dowmsampleBottleneck(channel_in, channel_out, stride=2):
    return nn.Sequential(
        nn.Conv2d(channel_in, 128, kernel_size=1, stride=1),
        nn.BatchNorm2d(128),
        nn.ReLU(),
        nn.Conv2d(128, 128, kernel_size=3, stride=stride, padding=1),
        nn.BatchNorm2d(128),
        nn.ReLU(),
        nn.Conv2d(128, channel_out, kernel_size=1, stride=1),
        nn.BatchNorm2d(channel_out),
        nn.ReLU(),
        )


class BasicBlock(nn.Module):
    """Basic Block for resnet 18 and resnet 34"""

    # BasicBlock and Bottleneck block
    # have different output size
    # we use class attribute expansion
    # to distinct
    expansion = 1

    def __init__(self, in_channels, out_channels, stride=1, use_conv=False):
        super().__init__()

        # residual function
        self.residual_function = nn.Sequential(
            nn.Conv2d(in_channels, out_channels, kernel_size=3, stride=stride, padding=1, bias=False),
            nn.BatchNorm2d(out_channels),
            nn.ReLU(inplace=True),
            nn.Conv2d(out_channels, out_channels * BasicBlock.expansion, kernel_size=3, padding=1, bias=False),
            nn.BatchNorm2d(out_channels * BasicBlock.expansion)
        )

        # shortcut
        self.shortcut = nn.Sequential()

        # the shortcut output dimension is not the same with residual function
        # use 1*1 convolution to match the dimension
        if stride != 1 or in_channels != out_channels * BasicBlock.expansion or use_conv:
            self.shortcut = nn.Sequential(
                nn.Conv2d(in_channels, out_channels * BasicBlock.expansion, kernel_size=1, stride=stride, bias=False),
                nn.BatchNorm2d(out_channels * BasicBlock.expansion)
            )
    @classmethod
    def from_digital(cls, digital_module, rpu_config, tile_module_class=None):
        # digital_module의 설정을 기반으로 아날로그 버전의 BasicBlock을 생성
        return cls(
            in_channels=digital_module.in_channels,
            out_channels=digital_module.out_channels,
            stride=digital_module.stride,
            use_conv=any(isinstance(layer, nn.Conv2d) for layer in digital_module.shortcut)
        )

    def forward(self, x):
        return nn.ReLU(inplace=True)(self.residual_function(x) + self.shortcut(x))


class Bottleneck(nn.Module):
    """Residual block for resnet over 50 layers"""

    expansion = 4

    def __init__(self, in_channels, out_channels, stride=1, use_conv=False):
        super().__init__()
        self.residual_function = nn.Sequential(
            nn.Conv2d(in_channels, out_channels, kernel_size=1, bias=False),
            nn.BatchNorm2d(out_channels),
            nn.ReLU(inplace=True),
            nn.Conv2d(out_channels, out_channels, kernel_size=3, stride=stride, padding=1, bias=False),
            nn.BatchNorm2d(out_channels),
            nn.ReLU(inplace=True),
            nn.Conv2d(out_channels, out_channels * Bottleneck.expansion, kernel_size=1, bias=False),
            nn.BatchNorm2d(out_channels * Bottleneck.expansion),
        )

        self.shortcut = nn.Sequential()

        if stride != 1 or in_channels != out_channels * Bottleneck.expansion or use_conv:
            self.shortcut = nn.Sequential(
                nn.Conv2d(in_channels, out_channels * Bottleneck.expansion, kernel_size=1, stride=stride, bias=False),
                nn.BatchNorm2d(out_channels * Bottleneck.expansion)
            )

    def forward(self, x):
        return nn.ReLU(inplace=True)(self.residual_function(x) + self.shortcut(x))

class ResNetBackbone(nn.Module):
    def __init__(self, block, layers, zero_init_residual=False, groups=1, 
                 width_per_group=64, replace_stride_with_dilation=None, norm_layer=None, num_classes=10):
        super(ResNetBackbone, self).__init__()
        self.block = block
        self.layers = layers
        self.zero_init_residual = zero_init_residual
        self.groups = groups
        self.base_width = width_per_group
        if replace_stride_with_dilation is None:
            self.replace_stride_with_dilation = [False, False, False]
        else:
            self.replace_stride_with_dilation = replace_stride_with_dilation
        self._norm_layer = norm_layer if norm_layer else nn.BatchNorm2d
        self.num_classes = num_classes
        self.inplanes = 64
        self.dilation = 1
        self.conv1 = nn.Conv2d(3, self.inplanes, kernel_size=3, stride=1, padding=1, bias=False)
        self.bn1 = self._norm_layer(self.inplanes)
        self.relu = nn.ReLU(inplace=True)
        self.layer1 = self._make_layer(block, 64, layers[0])
        self.layer2 = self._make_layer(block, 128, layers[1], stride=2, dilate=self.replace_stride_with_dilation[0])
        self.layer3 = self._make_layer(block, 256, layers[2], stride=2, dilate=self.replace_stride_with_dilation[1])
        self.layer4 = self._make_layer(block, 512, layers[3], stride=2, dilate=self.replace_stride_with_dilation[2])
        self.scala4 = nn.AvgPool2d(4, 4)
        self.fc4 = nn.Linear(512 * block.expansion, num_classes)

        for m in self.modules():
            if isinstance(m, nn.Conv2d):
                nn.init.kaiming_normal_(m.weight, mode='fan_out', nonlinearity='relu')
            elif isinstance(m, (nn.BatchNorm2d, nn.GroupNorm)):
                nn.init.constant_(m.weight, 1)
                nn.init.constant_(m.bias, 0)

        if zero_init_residual:
            for m in self.modules():
                if isinstance(m, Bottleneck):
                    nn.init.constant_(m.bn3.weight, 0)
                elif isinstance(m, BasicBlock):
                    nn.init.constant_(m.bn2.weight, 0)
    @classmethod
    def from_digital(cls, digital_module, rpu_config, tile_module_class=None):
        # 이 메서드는 digital_module (디지털 버전의 ResNetBackbone)을 기반으로 아날로그 버전을 생성합니다.
        # 필요에 따라 이 메서드를 세부적으로 수정할 수 있습니다.
        return cls(
            block=digital_module.block,
            layers=digital_module.layers,
            zero_init_residual=digital_module.zero_init_residual,
            groups=digital_module.groups,
            width_per_group=digital_module.base_width,
            replace_stride_with_dilation=digital_module.replace_stride_with_dilation,
            norm_layer=digital_module._norm_layer,
            num_classes=10  # 이것은 고정되어 있습니다. 필요한 경우 변경하십시오.
        )
    def _make_layer(self, block, planes, blocks, stride=1, dilate=False):
        norm_layer = self._norm_layer
        downsample = None
        previous_dilation = self.dilation
        if dilate:
            self.dilation *= stride
            stride = 1
        if stride != 1 or self.inplanes != planes * block.expansion:
            downsample = nn.Sequential(
                conv1x1(self.inplanes, planes * block.expansion, stride),
                norm_layer(planes * block.expansion),
            )

        layers = []
        use_conv = stride != 1 or self.inplanes != planes * block.expansion
        layers.append(block(self.inplanes, planes, stride, use_conv))
        self.inplanes = planes * block.expansion
        for _ in range(1, blocks):
            layers.append(block(self.inplanes, planes))

        return nn.Sequential(*layers)

    def forward(self, x):
        x = self.conv1(x)
        x = self.bn1(x)
        x = self.relu(x)
        x1 = self.layer1(x)  # x1 output
        x2 = self.layer2(x1)  # x2 output
        x3 = self.layer3(x2)  # x3 output
        x4 = self.layer4(x3)
        
        out4_feature = self.scala4(x4).view(x4.size(0), -1)
        if not hasattr(self, 'fc4') or self.fc4.in_features != out4_feature.size(1):
            self.fc4 = nn.Linear(out4_feature.size(1), 10).to(out4_feature.device)
            
        out4 = self.fc4(out4_feature)
        out4 = self.fc4(out4_feature)
        
        return out4, out4_feature, x1, x2, x3

class ResNetAttention1(nn.Module):
    def __init__(self, block, num_classes=100):  # num_classes 매개변수 추가
        super(ResNetAttention1, self).__init__()
        
        self.attention = nn.Sequential(
            SepConv(channel_in=64 * block.expansion, channel_out=64 * block.expansion),
            nn.BatchNorm2d(64 * block.expansion),
            nn.ReLU(),
            nn.Upsample(scale_factor=2, mode='bilinear'),
            nn.Sigmoid()
        )
        
        self.scala = nn.Sequential(
            SepConv(channel_in=64 * block.expansion, channel_out=128 * block.expansion),
            SepConv(channel_in=128 * block.expansion, channel_out=256 * block.expansion),
            SepConv(channel_in=256 * block.expansion, channel_out=512 * block.expansion),
            nn.AvgPool2d(4, 4)
        )
        
        self.fc = nn.Linear(512 * block.expansion, num_classes)  # 출력 차원을 num_classes로 설정

    def forward(self, x1):
        fea = self.attention(x1)
        fea = fea * x1
        feature_out = self.scala(fea).view(fea.size(0), -1)
        out = self.fc(feature_out)
        return out, feature_out
    
class ResNetAttention2(nn.Module):
    def __init__(self, block, num_classes=100):
        super(ResNetAttention2, self).__init__()
        
        self.attention = nn.Sequential(
            SepConv(channel_in=128 * block.expansion, channel_out=128 * block.expansion),
            nn.BatchNorm2d(128 * block.expansion),
            nn.ReLU(),
            nn.Upsample(scale_factor=2, mode='bilinear'),
            nn.Sigmoid()
        )
        
        self.scala = nn.Sequential(
            SepConv(channel_in=128 * block.expansion, channel_out=256 * block.expansion),
            SepConv(channel_in=256 * block.expansion, channel_out=512 * block.expansion),
            nn.AvgPool2d(4, 4)
        )
        
        self.fc = nn.Linear(512 * block.expansion, num_classes)

    def forward(self, x2):
        fea = self.attention(x2)
        fea = fea * x2
        feature_out = self.scala(fea).view(fea.size(0), -1)
        out = self.fc(feature_out)
        return out, feature_out


class ResNetAttention3(nn.Module):
    def __init__(self, block, num_classes=100):
        super(ResNetAttention3, self).__init__()
        
        self.attention = nn.Sequential(
            SepConv(channel_in=256 * block.expansion, channel_out=256 * block.expansion),
            nn.BatchNorm2d(256 * block.expansion),
            nn.ReLU(),
            nn.Upsample(scale_factor=2, mode='bilinear'),
            nn.Sigmoid()
        )
        
        self.scala = nn.Sequential(
            SepConv(channel_in=256 * block.expansion, channel_out=512 * block.expansion),
            nn.AvgPool2d(4, 4)
        )
        
        self.fc = nn.Linear(512 * block.expansion, num_classes)

    def forward(self, x3):
        fea = self.attention(x3)
        fea = fea * x3
        feature_out = self.scala(fea).view(fea.size(0), -1)
        out = self.fc(feature_out)
        return out, feature_out


def create_resnet(architecture="resnet34", num_classes=10):
    """Create a ResNet-inspired model using the ResNetBackbone class.

    Args:
        architecture (str): Which ResNet architecture to create (options: "resnet18", "resnet34", "resnet50")
        num_classes (int): Number of output classes

    Returns:
        nn.Module: Created model
    """
    if architecture == "resnet18":
        block = BasicBlock
        layers = [2, 2, 2, 2]
    elif architecture == "resnet34":
        block = BasicBlock
        layers = [3, 4, 6, 3]
    elif architecture == "resnet50":
        block = Bottleneck
        layers = [3, 4, 6, 3]
    elif architecture == "resnet10":
        block = BasicBlock
        layers = [1, 1, 1, 1]
            
    else:
        raise ValueError(f"Unknown architecture {architecture}")

    # Create the ResNetBackbone model
    model = ResNetBackbone(
        block=block,
        layers=layers,
        num_classes=num_classes,
        zero_init_residual=False,
        groups=1,
        width_per_group=64,
        replace_stride_with_dilation=None,
        norm_layer=nn.BatchNorm2d,
    )

    return model


def create_attention_resnets(num_classes=100, architecture="resnet34"):  # num_classes 매개변수 추가
    """Create ResNetAttention1, ResNetAttention2, and ResNetAttention3 models.

    Args:
        num_classes (int): Number of classes for the output layer. Default is 100.
        architecture (str): Which ResNet architecture to be used as a reference 
                            (options: "resnet18", "resnet34", "resnet50")

    Returns:
        tuple: Created ResNetAttention1, ResNetAttention2, ResNetAttention3 models
    """
    if architecture == "resnet18":
        block = BasicBlock
    elif architecture == "resnet34":
        block = BasicBlock
    elif architecture == "resnet50":
        block = Bottleneck
    elif architecture == "resnet10":
        block = BasicBlock
    else:
        raise ValueError(f"Unknown architecture {architecture}")

    # Create the ResNetAttention models
    model1 = ResNetAttention1(block=block, num_classes=num_classes)  # num_classes 매개변수 전달
    model2 = ResNetAttention2(block=block, num_classes=num_classes)  # num_classes 매개변수 전달
    model3 = ResNetAttention3(block=block, num_classes=num_classes)  # num_classes 매개변수 전달

    return model1, model2, model3



def create_sgd_optimizer(model, learning_rate):
    """Create the analog-aware optimizer.

    Args:
        model (nn.Module): model to be trained
        learning_rate (float): global parameter to define learning rate

    Returns:
        Optimizer: created analog optimizer
    """
    optimizer = AnalogSGD(model.parameters(), lr=learning_rate)
    optimizer.regroup_param_groups(model)

    return optimizer


# -*- coding: utf-8 -*-
# ... [중략: 기존 코드와 동일]
def run_model(model, images):
    return model(images)
def create_sgd_optimizer(model, learning_rate):
    """Create the analog-aware optimizer.

    Args:
        model (nn.Module): model to be trained
        learning_rate (float): global parameter to define learning rate

    Returns:
        Optimizer: created analog optimizer
    """
    optimizer = AnalogSGD(model.parameters(), lr=learning_rate)
    optimizer.regroup_param_groups(model)

    return optimizer

def train_step(train_data, model, criterion, optimizer):
    """Train network."""
    total_loss = 0

    model.train()

    for images, labels in train_data:
        images = images.to(DEVICE)
        labels = labels.to(DEVICE)
        optimizer.zero_grad()

        # Add training Tensor to the model (input).
        primary_output,_,_,_,_ = model(images)  # Handle multiple outputs
        loss = criterion(primary_output, labels)

        # Run training (backward propagation).
        loss.backward()

        # Optimize weights.
        optimizer.step()
        total_loss += loss.item() * images.size(0)
    epoch_loss = total_loss / len(train_data.dataset)

    return model, optimizer, epoch_loss

def test_evaluation(validation_data, model, criterion):
    """Test trained network"""
    total_loss = 0
    predicted_ok = 0
    total_images = 0

    model.eval()

    for images, labels in validation_data:
        images = images.to(DEVICE)
        labels = labels.to(DEVICE)

        primary_output,_,_,_,_= model(images)  # Handle multiple outputs
        loss = criterion(primary_output, labels)
        total_loss += loss.item() * images.size(0)

        _, predicted = torch_max(primary_output.data, 1)
        total_images += labels.size(0)
        predicted_ok += (predicted == labels).sum().item()
        accuracy = predicted_ok / total_images * 100
        error = (1 - predicted_ok / total_images) * 100

    epoch_loss = total_loss / len(validation_data.dataset)

    return model, epoch_loss, error, accuracy


def training_loop(model, criterion, optimizer, train_data, validation_data, epochs, print_every=1):
    """Training loop."""
    
    train_losses = []
    valid_losses = []
    test_error = []
    results_data = []  # List for saving epoch results

    # Train model
    for epoch in range(0, epochs):
        # Adjust learning rate based on the schedule
        if epoch in [epochs // 3, epochs * 2 // 3, epochs - 10]:
            for param_group in optimizer.param_groups:
                param_group["lr"] /= 10
        
        # Train_step
        model, optimizer, train_loss = train_step(train_data, model, criterion, optimizer)
        train_losses.append(train_loss)

        if epoch % print_every == (print_every - 1):
            # Validate_step
            with torch.no_grad():
                model, valid_loss, error, accuracy = test_evaluation(validation_data, model, criterion)
                valid_losses.append(valid_loss)
                test_error.append(error)

            print(
                f"{datetime.now().time().replace(microsecond=0)} --- "
                f"Epoch: {epoch}\t"
                f"Train loss: {train_loss:.4f}\t"
                f"Valid loss: {valid_loss:.4f}\t"
                f"Test error: {error:.2f}%\t"
                f"Test accuracy: {accuracy:.2f}%\t"
            )
            
            # Save the current epoch's results to results_data
            current_epoch_data = {
                "Epoch": epoch + 1,
                "Train Loss": train_loss,
                "Valid Loss": valid_loss,
                "Test Error": error,
                "Test Accuracy": accuracy
            }
            results_data.append(current_epoch_data)

    # Create and save the results as an Excel file
    df = pd.DataFrame(results_data)
    filepath = f"results/training_loop_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)

    return model, optimizer


def train_distill_step(inputs, labels, net_backbone, net_attention1, net_attention2, net_attention3, criterion, optimizer, device, args, adaptation_layers, init):
    inputs, labels = inputs.to(device), labels.to(device)

    # Zero the parameter gradients
    optimizer.zero_grad()

    # Forward pass
    # Obtain features using the backbone
    out_backbone, feature_backbone, x1, x2, x3 = net_backbone(inputs)

    # Obtain predictions using the attention mechanisms
    out_attention1, feature_attention1 = net_attention1(x1)
    out_attention2, feature_attention2 = net_attention2(x2)
    out_attention3, feature_attention3 = net_attention3(x3)

    outputs = [out_backbone, out_attention3, out_attention2, out_attention1]
    features = [feature_backbone, feature_attention3, feature_attention2, feature_attention1]

    if not init:
        # Initialize the adaptation layers
        teacher_feature_size = features[0].size(1)
        for feature in features[1:]:
            student_feature_size = feature.size(1)
            adaptation_layers.append(nn.Linear(student_feature_size, teacher_feature_size).to(device))
        init = True  # Make sure to return this value

    # Compute the self-distillation loss
    loss = criterion(outputs[0], labels)
    teacher_output = outputs[0].detach()
    teacher_feature = features[0].detach()

    for idx, (output, feature) in enumerate(zip(outputs[1:], features[1:])):
        # Logits distillation
        loss += CrossEntropy(output, teacher_output, args) * args.loss_coefficient
        loss += criterion(output, labels) * (1 - args.loss_coefficient)

        # Feature distillation
        if idx != 0:
            loss += (
                torch.dist(
                    adaptation_layers[idx-1](feature), teacher_feature
                )
                * args.feature_loss_coefficient
            )

    ensemble = sum(outputs) / len(outputs)
    outputs.append(ensemble.detach())  # Appending ensemble prediction to outputs list

    # Backward pass and optimize
    loss.backward()
    optimizer.step()

    # It might be necessary to return any values that are important for your training logic.
    # For instance, if you're tracking the loss or other metrics.
    return loss.item(), outputs, init  # return whatever values you need for the outer loop

# Train function for self-distillation
def train_distillation(trainloader, testloader, net_backbone, net_attention1, net_attention2, net_attention3, criterion, optimizer, device, args):
    init = False
    adaptation_layers = nn.ModuleList().to(device)
    results_data = []
    for epoch in range(args.epoch):
        net_backbone.train()
        net_attention1.train()
        net_attention2.train()
        net_attention3.train()

        if epoch in [args.epoch // 3, args.epoch * 2 // 3, args.epoch - 10]:
            for param_group in optimizer.param_groups:
                param_group["lr"] /= 10

        total_loss = 0
        total = 0
        correct = [0 for _ in range(5)]

        for i, (inputs, labels) in enumerate(trainloader):
            inputs, labels = inputs.to(device), labels.to(device)

            # Obtain features using the backbone
            out_backbone, feature_backbone, x1, x2, x3 = net_backbone(inputs)

            # Obtain predictions using the attention mechanisms
            out_attention1, feature_attention1 = net_attention1(x1)
            out_attention2, feature_attention2 = net_attention2(x2)
            out_attention3, feature_attention3 = net_attention3(x3)

            outputs = [out_backbone, out_attention3, out_attention2, out_attention1]
            features = [feature_backbone, feature_attention3, feature_attention2, feature_attention1]

            if not init:
                # Initialize the adaptation layers
                teacher_feature_size = features[0].size(1)
                for feature in features[1:]:
                    student_feature_size = feature.size(1)
                    adaptation_layers.append(nn.Linear(student_feature_size, teacher_feature_size).to(device))
                init = True

            # Compute the self-distillation loss
            loss = criterion(outputs[0], labels)
            teacher_output = outputs[0].detach()
            teacher_feature = features[0].detach()

            for idx, (output, feature) in enumerate(zip(outputs[1:], features[1:])):
                # Logits distillation
                loss += CrossEntropy(output, teacher_output, args) * args.loss_coefficient
                loss += criterion(output, labels) * (1 - args.loss_coefficient)

                # Feature distillation
                if idx != 0:
                    loss += (
                        torch.dist(
                            adaptation_layers[idx-1](feature), teacher_feature
                        )
                        * args.feature_loss_coefficient
                    )

            ensemble = sum(outputs) / len(outputs)
            outputs.append(ensemble.detach())  # Appending ensemble prediction to outputs list

            optimizer.zero_grad()
            loss.backward()
            optimizer.step()

             # Free up memory
            # del loss, out_backbone, out_attention1, out_attention2, out_attention3, outputs, features
            # torch.cuda.empty_cache()
            total_loss += loss.item()
            total += labels.size(0)
            
            for classifier_index in range(len(outputs)):
                _, predicted = torch.max(outputs[classifier_index].data, 1)
                correct[classifier_index] += float(predicted.eq(labels.data).cpu().sum())
            del loss, out_backbone, out_attention1, out_attention2, out_attention3, outputs, features
            torch.cuda.empty_cache()
        test_accuracies = evaluate_distillation(testloader, net_backbone, net_attention1, net_attention2, net_attention3, device)
        current_epoch_data = {
            "Epoch": epoch + 1,
            "Train Loss": total_loss / (i + 1),
            "Test Acc Backbone": test_accuracies[0],
            "Test Acc Att3": test_accuracies[1],
            "Test Acc Att2": test_accuracies[2],
            "Test Acc Att1": test_accuracies[3],
            "Test Acc Ensemble": test_accuracies[4]
            }
        results_data.append(current_epoch_data)        
        # 현재 에포크
    df = pd.DataFrame(results_data)
    filepath = f"results/results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
        
    return

def train_distillation_loop(trainloader, testloader, net_backbone, net_attention1, net_attention2, net_attention3, criterion, optimizer, device, args):
    # 초기 설정
    init = False
    adaptation_layers = nn.ModuleList().to(device)
    results_data = []

    for epoch in range(args.epoch):
        net_backbone.train()
        net_attention1.train()
        net_attention2.train()
        net_attention3.train()

        # Learning rate 조정
        if epoch in [args.epoch // 3, args.epoch * 2 // 3, args.epoch - 10]:
            for param_group in optimizer.param_groups:
                param_group["lr"] /= 10

        total_loss = 0
        total = 0
        correct = [0 for _ in range(5)]  # 각 모델의 정확도를 추적

        for i, (inputs, labels) in enumerate(trainloader):
            # train_distill_step 함수 호출
            loss = train_distill_step(inputs, labels, net_backbone, net_attention1, net_attention2, net_attention3, criterion, optimizer, device, args, adaptation_layers, init)
            total_loss += loss
            total += labels.size(0)

            # 정확도 계산
            outputs = net_backbone(inputs)  # 예측값 가져오기
            _, predicted = torch.max(outputs, 1)
            correct += (predicted == labels).sum().item()

        # 에포크당 평균 손실 계산
        average_loss = total_loss / len(trainloader)
        accuracy = 100 * correct / total

        # 결과 저장
        results_data.append({
            "Epoch": epoch + 1,
            "Loss": average_loss,
            "Accuracy": accuracy
        })

        # 테스트 데이터로 모델 평가
        test_accuracies = evaluate_distillation(testloader, net_backbone, net_attention1, net_attention2, net_attention3, device)

        # 결과 출력
        print(f'Epoch {epoch+1}, Loss: {average_loss:.4f}, Accuracy: {accuracy:.2f}%')
        print('Test Accuracies:', test_accuracies)

    # 결과를 Excel 파일로 저장
    df = pd.DataFrame(results_data)
    filepath = f"results/results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)

    return net_backbone, net_attention1, net_attention2, net_attention3

def evaluate_distillation(testloader, net_backbone, net_attention1, net_attention2, net_attention3, device):
    net_backbone.eval()
    net_attention1.eval()
    net_attention2.eval()
    net_attention3.eval()

    correct = [0 for _ in range(5)]
    total = 0

    with torch.no_grad():
        for data in testloader:
            images, labels = data
            images, labels = images.to(device), labels.to(device)

            # Obtain features and outputs using the backbone and attention mechanisms
            out_backbone, _, x1, x2, x3 = net_backbone(images)
            out_attention1, _ = net_attention1(x1)
            out_attention2, _ = net_attention2(x2)
            out_attention3, _ = net_attention3(x3)

            outputs = [out_backbone, out_attention3, out_attention2, out_attention1]
            
            ensemble = sum(outputs) / len(outputs)
            outputs.append(ensemble)  # Appending ensemble prediction to outputs list

            for classifier_index in range(len(outputs)):
                _, predicted = torch.max(outputs[classifier_index].data, 1)
                correct[classifier_index] += float(predicted.eq(labels.data).cpu().sum())
            total += float(labels.size(0))

    print('Test Set Accuracy Acc: 4/4: %.4f%% 3/4: %.4f%% 2/4: %.4f%% 1/4: %.4f%% Ensemble: %.4f%%' 
          % (100 * correct[0] / total, 100 * correct[1] / total, 100 * correct[2] / total, 
             100 * correct[3] / total, 100 * correct[4] / total))

    return [100. * c / total for c in correct]

import pandas as pd
from datetime import datetime

def apply_salmon(epochs, distill_epoch, trainloader, testloader, net_backbone, net_attention1, net_attention2, net_attention3, criterion, optimizer, device, args):
    adaptation_layers = nn.ModuleList().to(device)
    init = False  # Adaptation layers' initialization flag

    # 결과를 저장할 리스트 생성
    results_data = []

    for epoch in range(epochs):
        # 학습률 조정 로직
        if epoch in [epochs // 3, epochs * 2 // 3, epochs - 10]:
            for param_group in optimizer.param_groups:
                param_group["lr"] /= 10

        if epoch > 0 and epoch % distill_epoch == 0:  # Check if it's time for distillation
            print(f"Epoch {epoch}: Starting distillation...")
            net_backbone.train()
            net_attention1.train()
            net_attention2.train()
            net_attention3.train()

            for inputs, labels in trainloader:
                # Distillation step
                _, _, init = train_distill_step(inputs, labels, net_backbone, net_attention1, net_attention2, net_attention3, criterion, optimizer, device, args, adaptation_layers, init)

            # After distillation, evaluate the model
            print("Evaluating after distillation...")
            accuracies = evaluate_distillation(testloader, net_backbone, net_attention1, net_attention2, net_attention3, device)
            
            # 정확도를 결과 데이터에 추가
            results_data.append({
                "Epoch": epoch, 
                "Backbone Accuracy": accuracies[0], 
                "Attention1 Accuracy": accuracies[1], 
                "Attention2 Accuracy": accuracies[2], 
                "Attention3 Accuracy": accuracies[3], 
                "Ensemble Accuracy": accuracies[4]
            })

        else:  # Regular training
            print(f"Epoch {epoch}: Starting regular training...")
            # Train the model regularly
            net_backbone, optimizer, epoch_loss = train_step(trainloader, net_backbone, criterion, optimizer)
            print(f"Epoch {epoch} loss: {epoch_loss}")

            # Evaluate your model after regular training
            print("Evaluating after regular training...")
            _, test_loss, error, accuracy = test_evaluation(testloader, net_backbone, criterion)  # Assuming these functions return these values
            print(f"Test loss: {test_loss}, Accuracy: {accuracy}")

            # 정확도를 결과 데이터에 추가
            results_data.append({
                "Epoch": epoch, 
                "Backbone Accuracy": accuracy, 
                # 다른 attention 모델의 정확도도 같은 방식으로 추가할 수 있습니다.
            })

    # 훈련이 끝난 후 결과를 Excel 파일로 저장
    df = pd.DataFrame(results_data)
    
    # 파일 이름에 distill_epoch 값을 포함시킵니다.
    current_time = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"salmon_{distill_epoch}_{current_time}.xlsx"
    filepath = os.path.join("results", filename)  # 'results' 디렉토리에 파일 저장

    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)

    return net_backbone, net_attention1, net_attention2, net_attention3  # Return models if you plan to use them after training

# 기본적인 KD train module

def distillation_loss(student_outputs, teacher_outputs, labels, temperature, alpha):

    soft_loss = F.kl_div(
        F.log_softmax(student_outputs / temperature, dim=1),
        F.softmax(teacher_outputs / temperature, dim=1),
        reduction='batchmean',
    ) * (temperature * temperature)
    
    hard_loss = F.cross_entropy(student_outputs, labels)
    
    return alpha * hard_loss + (1. - alpha) * soft_loss

def train_distillation(student_model, teacher_model, train_loader, optimizer, temperature, alpha):
    student_model.train()
    teacher_model.eval()
    
    total_loss = 0.0
    
    for images, labels in train_loader:
        images, labels = images.to(DEVICE), labels.to(DEVICE)
        
        # images = images.view(images.shape[0], -1)  # 이 줄을 제거하세요
        
        optimizer.zero_grad()
        student_outputs = student_model(images)
        with torch.no_grad():
            teacher_outputs = teacher_model(images)
        
        loss = distillation_loss(student_outputs, teacher_outputs, labels, temperature, alpha)
        loss.backward()
        optimizer.step()
        
        total_loss += loss.item()

    average_loss = total_loss / len(train_loader)
    
    return average_loss  # Return the average loss for this epoch


def training_loop_distillation(student_model, teacher_model, criterion, optimizer, train_data, validation_data, temperature, alpha, epochs, distill_every=1, print_every=1):
    train_losses = []
    valid_losses = []
    test_error = []
    accuracies = []  # To store accuracies for each epoch
    
    # Train model
    for epoch in range(epochs):
        
        # Learning rate adjustment logic
        if epoch in [epochs // 3, epochs * 2 // 3, epochs - 10]:
            for param_group in optimizer.param_groups:
                param_group["lr"] /= 10
                
        if epoch % distill_every == 0:  # If the current epoch is a multiple of distill_every, perform distillation
            average_loss = train_distillation(student_model, teacher_model, train_data, optimizer, temperature, alpha)
        else:  # Otherwise, perform standard training
            student_model, optimizer, average_loss = train_step(train_data, student_model, criterion, optimizer)
            
        train_losses.append(average_loss)
        
        # Validation
        if epoch % print_every == (print_every - 1):
            with torch.no_grad():
                student_model.eval()  # Set student_model to evaluation mode
                _, valid_loss, error, accuracy = test_evaluation(validation_data, student_model, criterion)
                valid_losses.append(valid_loss)
                test_error.append(error)
                accuracies.append(accuracy)  # Store the accuracy for this epoch

            print(f"{datetime.now().time().replace(microsecond=0)} --- "
                  f"Epoch: {epoch}\t"
                  f"Train loss: {average_loss:.4f}\t"
                  f"Valid loss: {valid_loss:.4f}\t"
                  f"Test error: {error:.2f}%\t"
                  f"Test accuracy: {accuracy:.2f}%\t")
    
    # plot_results(train_losses, valid_losses, test_error)

    # Convert results to pandas DataFrame
    results_data = {
        "Train Losses": train_losses,
        "Validation Losses": valid_losses,
        "Test Error": test_error,
        "Accuracies": accuracies
    }
    df = pd.DataFrame(results_data)
    
    # Save DataFrame to Excel
    current_time = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"kd_results_{distill_every}_{current_time}.xlsx"
    filepath = os.path.join("results", filename)  # 'results' directory to store the file
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)

    return student_model, optimizer, (train_losses, valid_losses, test_error, accuracies)
